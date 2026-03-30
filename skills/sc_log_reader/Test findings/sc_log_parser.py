"""
SC Game.log Parser - Parses Star Citizen Game.log files into structured Excel spreadsheets.
Developer: Mallachi
Version: 1.1.0
"""

import re
import sys
import glob
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

VERSION = "1.1.0"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

TIMESTAMP_RE = re.compile(r"^<(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d+Z)>\s*(.*)")
EVENT_TYPE_RE = re.compile(r"^\[([^\]]+)\]\s*(.*)")
SUBTYPE_RE = re.compile(r"^<([^>]+)>\s*(.*)")
ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGN = Alignment(horizontal="left", vertical="top")
CELL_ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ROW_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def sanitize(text):
    if not text:
        return text
    text = ILLEGAL_CHARS_RE.sub("", text)
    if len(text) > 32767:
        text = text[:32760] + "...[truncated]"
    return text


def parse_log_file(filepath):
    """Parse a Game.log file into [timestamp, event_type, sub_type, content] entries."""
    entries = []
    current_entry = None

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\r\n")
            m = TIMESTAMP_RE.match(line)
            if m:
                if current_entry:
                    entries.append(current_entry)
                timestamp_str = m.group(1)
                rest = m.group(2)

                event_type = ""
                sub_type = ""

                # Check for [EventType]
                em = EVENT_TYPE_RE.match(rest)
                if em:
                    event_type = em.group(1)
                    rest = em.group(2)

                # Check for <SubType>
                sm = SUBTYPE_RE.match(rest)
                if sm:
                    sub_type = sm.group(1)
                    rest = sm.group(2)

                current_entry = [timestamp_str, event_type, sub_type, rest]
            else:
                if current_entry:
                    current_entry[3] += "\n" + line

    if current_entry:
        entries.append(current_entry)
    return entries


def source_name(filepath):
    name = os.path.splitext(os.path.basename(filepath))[0]
    for suffix in ["_Game", "_game"]:
        if name.endswith(suffix):
            name = name[: -len(suffix)]
    return name


def style_header_row(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def main():
    log_files = sorted(glob.glob(os.path.join(SCRIPT_DIR, "*.log")))
    if not log_files:
        print(f"No .log files found in {SCRIPT_DIR}")
        sys.exit(0)

    print(f"SC Game.log Parser v{VERSION}")
    print(f"Directory: {SCRIPT_DIR}")
    print(f"Found {len(log_files)} log file(s)\n")

    parsed = {}
    total_entries = 0
    for fp in log_files:
        name = source_name(fp)
        print(f"  Parsing {os.path.basename(fp)}...")
        entries = parse_log_file(fp)
        parsed[name] = entries
        total_entries += len(entries)
        print(f"    {len(entries):,} entries")

    print(f"\n  Total: {total_entries:,} entries across {len(log_files)} files\n")

    wb = Workbook()
    wb.remove(wb.active)

    # Combined sheet
    if len(parsed) > 1:
        print("  Building combined sheet...")
        ws_all = wb.create_sheet(title="All Logs", index=0)
        ws_all.append(["Source", "Timestamp", "Event Type", "Sub-Type", "Content"])
        style_header_row(ws_all, 5)

        row_idx = 2
        for name, entries in parsed.items():
            for ts, etype, stype, content in entries:
                ws_all.cell(row=row_idx, column=1, value=name).font = CELL_FONT
                ws_all.cell(row=row_idx, column=2, value=ts).font = CELL_FONT
                ws_all.cell(row=row_idx, column=2).alignment = CELL_ALIGN
                ws_all.cell(
                    row=row_idx, column=3, value=sanitize(etype)
                ).font = CELL_FONT
                ws_all.cell(
                    row=row_idx, column=4, value=sanitize(stype)
                ).font = CELL_FONT
                ws_all.cell(
                    row=row_idx, column=5, value=sanitize(content)
                ).font = CELL_FONT
                ws_all.cell(row=row_idx, column=5).alignment = CELL_ALIGN_WRAP
                row_idx += 1

        ws_all.freeze_panes = "A2"
        ws_all.column_dimensions["A"].width = 18
        ws_all.column_dimensions["B"].width = 28
        ws_all.column_dimensions["C"].width = 16
        ws_all.column_dimensions["D"].width = 45
        ws_all.column_dimensions["E"].width = 120

    # Individual sheets
    for name, entries in parsed.items():
        print(f"  Building sheet: {name}")
        sheet_name = name[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Timestamp", "Event Type", "Sub-Type", "Content"])
        style_header_row(ws, 4)

        for i, (ts, etype, stype, content) in enumerate(entries, start=2):
            ws.cell(row=i, column=1, value=ts).font = CELL_FONT
            ws.cell(row=i, column=1).alignment = CELL_ALIGN
            ws.cell(row=i, column=2, value=sanitize(etype)).font = CELL_FONT
            ws.cell(row=i, column=3, value=sanitize(stype)).font = CELL_FONT
            ws.cell(row=i, column=4, value=sanitize(content)).font = CELL_FONT
            ws.cell(row=i, column=4).alignment = CELL_ALIGN_WRAP

        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 120

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(SCRIPT_DIR, f"GameLogs_Parsed_{timestamp}.xlsx")
    print(f"\n  Saving to {os.path.basename(outfile)}...")
    wb.save(outfile)
    print(f"  Done. {total_entries:,} entries written.")


if __name__ == "__main__":
    main()
